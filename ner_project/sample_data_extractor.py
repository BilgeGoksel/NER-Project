from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook
import threading
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import os

# --- Config ---
chrome_driver_path = r"C:\Users\stj.skartal\Desktop\chromedriver-win64\chromedriver.exe"

# ÇIKTI klasörü ve dosyası -> data/sikayetvar.xlsx
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
excel_file = os.path.join(DATA_DIR, "sikayetvar.xlsx")

lock = threading.Lock()

# (Önceden 'paraler' idi) -> şirket listesi
sirketler = [
    "Koç", "Sabancı", "TÜPRAŞ", "BİM", "ASELSAN", "Arçelik", "Vestel", "THY", "Türk Telekom", "Turkcell","LC Waikiki", "Migros", "Ford", "TOFAŞ", "Anadolu Efes", "Şişecam", "Eczacıbaşı", "Ziraat", "Garanti","Akbank", "İş Bankası", "Yapı Kredi", "Halkbank", "QNB", "VakıfBank", "Opet", "Petrol Ofisi","Enka", "Doğuş", "Çalık", "Limak", "Rönesans", "MNG", "Trendyol", "Hepsiburada", "Getir", "Yemeksepeti","Pegasus", "SunExpress", "Baykar", "Roketsan", "Havelsan", "Otokar", "BMC", "Temsa", "Koton", "Defacto","Penti", "Mavi", "FLO", "Vatan", "Teknosa", "Sahibinden", "N11", "Morhipo", "Boyner", "Watsons","Gratis", "A101", "Şok", "Metro", "CarrefourSA", "LCW", "Zorlu", "Doğan", "Alarko", "Borusan","Eren", "Demirören", "Cengiz", "Kolin", "Nurol", "OYAK", "Kale", "TAV", "İGA", "Sabiha","Eti", "Ülker", "Pınar", "Sütaş", "Torku", "Banvit", "Namet", "Tat", "Dardanel", "Arzum","Fakir", "Korkmaz", "Simfer", "Vitra", "Artema", "Bosch", "Siemens", "Mercedes", "MAN"
]

# --- Excel Yardımcıları ---
def init_excel():
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "Şirket", "Başlık", "Link", "Şikayet", "Çekilme Tarihi"])
        wb.save(excel_file)

def append_to_excel(data):
    with lock:
        if not os.path.exists(excel_file):
            init_excel()
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([
            data["ID"],
            data["Şirket"],
            data["Başlık"],
            data["Link"],
            data["Şikayet"],
            data["Çekilme Tarihi"]
        ])
        wb.save(excel_file)
        wb.close()

def get_existing_titles():
    titles = set()
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            title = row[2]  # 'Başlık' sütunu
            if title:
                titles.add(title)
        wb.close()
    return titles

# --- Selenium ---
def create_driver():
    options = Options()
    options.add_argument("--headless=new")  # yeni headless
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--log-level=3")
    service = Service(chrome_driver_path)
    return webdriver.Chrome(service=service, options=options)

def fetch_complaint(url, sirket, title, idx):
    driver = create_driver()
    try:
        driver.get(url)
        time.sleep(2)
        try:
            complaint_div = driver.find_element(By.CSS_SELECTOR, "div.complaint-detail-description")
            complaint_text = complaint_div.text.strip()
        except:
            complaint_text = "Şikayet metni bulunamadı."

        result = {
            "ID": idx,
            "Şirket": sirket,
            "Başlık": title,
            "Link": url,
            "Şikayet": complaint_text,
            "Çekilme Tarihi": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        append_to_excel(result)
        print(f" Yazıldı: {title}")
    except Exception as e:
        result = {
            "ID": idx,
            "Şirket": sirket,
            "Başlık": title,
            "Link": url,
            "Şikayet": f"Hata: {str(e)}",
            "Çekilme Tarihi": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        append_to_excel(result)
        print(f" Hata yazıldı: {title}")
    finally:
        driver.quit()
    return result

# --- Ana Akış ---
def main():
    init_excel()
    idx = 0
    yazilan_basliklar = get_existing_titles()

    print("🔍 Şirketlere göre şikayetler toplanıyor...")

    with ThreadPoolExecutor(max_workers=3) as executor:
        futures = []

        for sirket in sirketler:
            print(f"\n Şirket: {sirket}")
            driver = create_driver()
            try:
                for page in range(1, 5):
                    url = f"https://www.sikayetvar.com/sikayetler?k={sirket}&page={page}"
                    print(f"   Sayfa {page}: {url}")
                    driver.get(url)
                    time.sleep(3)
                    links = driver.find_elements(By.CSS_SELECTOR, "h2.complaint-title a")
                    if not links:
                        print("     Link bulunamadı, sonraki şirkete geçiliyor.")
                        break

                    for link in links:
                        title = (link.text or "").strip()
                        if not title:
                            continue
                        if title in yazilan_basliklar:
                            print(f"     '{title}' başlığı zaten yazılmış, atlanıyor.")
                            continue

                        idx += 1
                        href = link.get_attribute("href")
                        yazilan_basliklar.add(title)
                        futures.append(executor.submit(fetch_complaint, href, sirket, title, idx))
            except Exception as e:
                print(f"   Hata: {e}")
            finally:
                driver.quit()

        # Detay görevlerini bitir
        for future in as_completed(futures):
            future.result()

    print(f"\n Veri çekme işlemi tamamlandı. Excel yolu: {excel_file}")

if __name__ == "__main__":
    main()

